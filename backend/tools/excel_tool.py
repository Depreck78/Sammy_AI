import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

from app.config import APP_ROOT
from app.tooling import BaseTool


WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
DEFAULT_OUTPUT_ROOT = Path.home() / ".sammy" / "outputs" / "spreadsheets"


def _openpyxl() -> Tuple[Any, Any, Any, Any]:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
    except ImportError as exc:
        raise RuntimeError("openpyxl is not installed. Run pip install -r requirements.txt and restart Sammy.") from exc
    return Workbook, load_workbook, coordinate_to_tuple, range_boundaries


def _json(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, indent=2, default=str)


class ExcelTool(BaseTool):
    name = "excel"
    display_name = "Excel"
    description = (
        "Inspect, read, search, and create local Excel .xlsx workbooks. New workbooks can be "
        "created in Sammy's safe output folder even when broader spreadsheet writes are disabled."
    )
    icon = "FileSpreadsheet"
    requires_auth = False

    def get_auth_fields(self) -> List[Dict[str, Any]]:
        return [
            {
                "name": "allowed_directories",
                "label": "Allowed directories",
                "type": "textarea",
                "placeholder": str(APP_ROOT),
                "description": "One absolute directory path per line. Defaults to this Sammy repository only.",
            },
            {
                "name": "allow_writes",
                "label": "Allow spreadsheet writes",
                "type": "checkbox",
                "description": (
                    "Off by default. Sammy can still create new workbooks in its own safe output folder; "
                    "enable this only for other folders where Sammy may create or edit workbooks."
                ),
            },
        ]

    def _allowed_roots(self) -> List[Path]:
        raw = self.credentials.get("allowed_directories")
        if isinstance(raw, str):
            values = [line.strip() for line in raw.splitlines() if line.strip()]
        elif isinstance(raw, list):
            values = raw
        else:
            values = [str(APP_ROOT)]
        roots = []
        for value in values:
            try:
                roots.append(Path(value).expanduser().resolve())
            except Exception:
                continue
        return roots or [APP_ROOT.resolve()]

    def _writes_enabled(self) -> bool:
        value = self.credentials.get("allow_writes")
        if isinstance(value, bool):
            return value
        return str(value or "").strip().lower() in {"1", "true", "yes", "on"}

    def _resolve_allowed(self, path: str) -> Path:
        candidate = Path(path).expanduser()
        if not candidate.is_absolute():
            candidate = APP_ROOT / candidate
        candidate = candidate.resolve()
        if candidate.suffix.lower() not in WORKBOOK_EXTENSIONS:
            raise ValueError("Excel tool supports .xlsx, .xlsm, .xltx, and .xltm files.")
        safe_root = self._safe_output_root()
        if candidate == safe_root or safe_root in candidate.parents:
            return candidate
        for root in self._allowed_roots():
            if candidate == root or root in candidate.parents:
                return candidate
        allowed = ", ".join(str(root) for root in self._allowed_roots())
        raise PermissionError(f"Path is outside allowed directories. Allowed: {allowed}")

    def _safe_output_root(self) -> Path:
        return DEFAULT_OUTPUT_ROOT.expanduser().resolve()

    def _slug(self, value: str) -> str:
        slug = re.sub(r"[^a-zA-Z0-9]+", "_", str(value or "").strip()).strip("_").lower()
        return slug[:48] or "excel_sheet"

    def _default_output_path(self, parameters: Dict[str, Any]) -> Path:
        label = parameters.get("title") or parameters.get("sheet_name") or "excel_sheet"
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return self._safe_output_root() / f"{self._slug(str(label))}_{stamp}.xlsx"

    def _resolve_create_path(self, parameters: Dict[str, Any]) -> Path:
        raw_path = str(parameters.get("path") or "").strip()
        if not raw_path:
            return self._default_output_path(parameters)
        path = Path(raw_path).expanduser()
        if path.suffix == "":
            path = path / f"{self._slug(str(parameters.get('title') or parameters.get('sheet_name') or 'excel_sheet'))}.xlsx"
        if not path.is_absolute():
            path = APP_ROOT / path
        path = path.resolve()
        if path.suffix.lower() not in WORKBOOK_EXTENSIONS:
            raise ValueError("Excel tool supports .xlsx, .xlsm, .xltx, and .xltm files.")
        safe_root = self._safe_output_root()
        if path == safe_root or safe_root in path.parents:
            return path
        if not self._writes_enabled():
            raise PermissionError(
                f"Creating outside Sammy's safe spreadsheet folder requires Allow spreadsheet writes. "
                f"Safe folder: {safe_root}"
            )
        return self._resolve_allowed(str(path))

    def _load_workbook(self, path: Path, data_only: bool = True) -> Any:
        _, load_workbook, _, _ = _openpyxl()
        if not path.exists():
            raise FileNotFoundError(path)
        return load_workbook(path, data_only=data_only)

    def _sheet(self, workbook: Any, sheet_name: str = "") -> Any:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet not found: {sheet_name}")
            return workbook[sheet_name]
        return workbook.active

    def _coerce_rows(self, value: Any) -> List[List[Any]]:
        if value is None:
            return []
        if isinstance(value, str):
            value = json.loads(value)
        if not isinstance(value, list):
            raise ValueError("Rows must be an array of row arrays.")
        rows: List[List[Any]] = []
        for row in value:
            if isinstance(row, list):
                rows.append(row)
            elif isinstance(row, dict):
                rows.append(list(row.values()))
            else:
                rows.append([row])
        return rows

    def _coerce_headers(self, value: Any) -> List[Any]:
        if value is None or value == "":
            return []
        if isinstance(value, str):
            try:
                value = json.loads(value)
            except json.JSONDecodeError:
                value = [part.strip() for part in value.split(",") if part.strip()]
        if not isinstance(value, list):
            raise ValueError("Headers must be an array or comma-separated string.")
        return value

    def get_functions(self) -> List[Dict[str, Any]]:
        functions = [
            self.function(
                "excel_inspect_workbook",
                "Inspect an Excel workbook: sheet names, dimensions, and a small preview.",
                {
                    "path": {"type": "string", "description": "Absolute or workspace-relative .xlsx path."},
                    "preview_rows": {"type": "integer", "minimum": 1, "maximum": 20},
                },
                ["path"],
            ),
            self.function(
                "excel_read_range",
                "Read a range from an Excel workbook as JSON rows.",
                {
                    "path": {"type": "string"},
                    "sheet_name": {"type": "string", "description": "Optional. Defaults to the active sheet."},
                    "range": {"type": "string", "description": "Optional Excel range such as A1:D20. Defaults to used rows."},
                    "max_rows": {"type": "integer", "minimum": 1, "maximum": 500},
                    "max_columns": {"type": "integer", "minimum": 1, "maximum": 100},
                },
                ["path"],
            ),
            self.function(
                "excel_search_workbook",
                "Search visible cell values in an Excel workbook.",
                {
                    "path": {"type": "string"},
                    "query": {"type": "string"},
                    "max_matches": {"type": "integer", "minimum": 1, "maximum": 100},
                },
                ["path", "query"],
            ),
            self.function(
                "excel_create_workbook",
                (
                    "Create a new Excel .xlsx workbook with optional headers and rows. The path is optional; "
                    "if omitted, Sammy saves the file in its safe spreadsheet output folder. Use this for user "
                    "requests like creating a table or making an Excel sheet."
                ),
                {
                    "path": {
                        "type": "string",
                        "description": "Optional. Absolute or workspace-relative .xlsx path. Omit to use Sammy's safe spreadsheet output folder.",
                    },
                    "title": {"type": "string", "description": "Optional filename hint, such as colors table."},
                    "sheet_name": {"type": "string"},
                    "headers": {"type": "array", "items": {}},
                    "rows": {"type": "array", "items": {"type": "array", "items": {}}},
                },
            ),
        ]
        if self._writes_enabled():
            functions.extend(
                [
                    self.function(
                        "excel_write_range",
                        "Write rows into an Excel workbook starting at a cell. Creates the workbook or sheet if needed. Requires Allow spreadsheet writes.",
                        {
                            "path": {"type": "string"},
                            "sheet_name": {"type": "string"},
                            "start_cell": {"type": "string", "description": "Top-left cell such as A1."},
                            "values": {"type": "array", "items": {"type": "array", "items": {}}},
                        },
                        ["path", "start_cell", "values"],
                    ),
                    self.function(
                        "excel_append_rows",
                        "Append rows to the end of a sheet in an Excel workbook. Requires Allow spreadsheet writes.",
                        {
                            "path": {"type": "string"},
                            "sheet_name": {"type": "string"},
                            "rows": {"type": "array", "items": {"type": "array", "items": {}}},
                        },
                        ["path", "rows"],
                    ),
                ]
            )
        return functions

    def _read_rows(self, worksheet: Any, range_ref: str, max_rows: int, max_columns: int) -> Dict[str, Any]:
        _, _, _, range_boundaries = _openpyxl()
        max_rows = max(1, min(int(max_rows or 50), 500))
        max_columns = max(1, min(int(max_columns or 50), 100))
        if range_ref:
            min_col, min_row, max_col, max_row = range_boundaries(range_ref)
            max_row = min(max_row, min_row + max_rows - 1)
            max_col = min(max_col, min_col + max_columns - 1)
        else:
            min_col, min_row = 1, 1
            max_row = min(worksheet.max_row or 1, max_rows)
            max_col = min(worksheet.max_column or 1, max_columns)
        rows = [
            [worksheet.cell(row=row, column=col).value for col in range(min_col, max_col + 1)]
            for row in range(min_row, max_row + 1)
        ]
        return {
            "sheet": worksheet.title,
            "range": f"{worksheet.cell(min_row, min_col).coordinate}:{worksheet.cell(max_row, max_col).coordinate}",
            "rows": rows,
        }

    def _inspect(self, parameters: Dict[str, Any]) -> str:
        path = self._resolve_allowed(parameters["path"])
        workbook = self._load_workbook(path)
        preview_rows = max(1, min(int(parameters.get("preview_rows") or 5), 20))
        sheets = []
        for worksheet in workbook.worksheets:
            preview = self._read_rows(worksheet, "", preview_rows, 12)["rows"]
            sheets.append(
                {
                    "name": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "preview": preview,
                }
            )
        return _json({"path": str(path), "sheets": sheets})

    def _read_range(self, parameters: Dict[str, Any]) -> str:
        path = self._resolve_allowed(parameters["path"])
        workbook = self._load_workbook(path)
        worksheet = self._sheet(workbook, parameters.get("sheet_name") or "")
        return _json(
            self._read_rows(
                worksheet,
                parameters.get("range") or "",
                int(parameters.get("max_rows") or 50),
                int(parameters.get("max_columns") or 50),
            )
        )

    def _search(self, parameters: Dict[str, Any]) -> str:
        path = self._resolve_allowed(parameters["path"])
        query = str(parameters.get("query") or "").casefold()
        if not query:
            raise ValueError("Search query cannot be empty.")
        max_matches = max(1, min(int(parameters.get("max_matches") or 50), 100))
        workbook = self._load_workbook(path)
        matches = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is not None and query in str(value).casefold():
                        matches.append({"sheet": worksheet.title, "cell": cell.coordinate, "value": value})
                        if len(matches) >= max_matches:
                            return _json({"matches": matches})
        return _json({"matches": matches})

    def _create(self, parameters: Dict[str, Any]) -> str:
        Workbook, _, _, _ = _openpyxl()
        path = self._resolve_create_path(parameters)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = parameters.get("sheet_name") or "Sheet1"
        headers = self._coerce_headers(parameters.get("headers"))
        rows = self._coerce_rows(parameters.get("rows"))
        if headers:
            worksheet.append(headers)
        for row in rows:
            worksheet.append(row)
        workbook.save(path)
        return f"Created Excel workbook: {path}"

    def _write_range(self, parameters: Dict[str, Any]) -> str:
        if not self._writes_enabled():
            return "Excel writes are disabled. Enable Allow spreadsheet writes in the Excel tool settings first."
        Workbook, load_workbook, coordinate_to_tuple, _ = _openpyxl()
        path = self._resolve_allowed(parameters["path"])
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = load_workbook(path) if path.exists() else Workbook()
        sheet_name = parameters.get("sheet_name") or workbook.active.title
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
        start_row, start_col = coordinate_to_tuple(parameters.get("start_cell") or "A1")
        rows = self._coerce_rows(parameters.get("values"))
        for row_offset, row in enumerate(rows):
            for col_offset, value in enumerate(row):
                worksheet.cell(row=start_row + row_offset, column=start_col + col_offset, value=value)
        workbook.save(path)
        return f"Wrote {len(rows)} rows to {worksheet.title} starting at {parameters.get('start_cell') or 'A1'} in {path}"

    def _append_rows(self, parameters: Dict[str, Any]) -> str:
        if not self._writes_enabled():
            return "Excel writes are disabled. Enable Allow spreadsheet writes in the Excel tool settings first."
        _, load_workbook, _, _ = _openpyxl()
        path = self._resolve_allowed(parameters["path"])
        if not path.exists():
            raise FileNotFoundError(path)
        workbook = load_workbook(path)
        worksheet = self._sheet(workbook, parameters.get("sheet_name") or "")
        rows = self._coerce_rows(parameters.get("rows"))
        for row in rows:
            worksheet.append(row)
        workbook.save(path)
        return f"Appended {len(rows)} rows to {worksheet.title} in {path}"

    def execute(self, function_name: str, parameters: Dict[str, Any]) -> str:
        try:
            if function_name == "excel_inspect_workbook":
                return self._inspect(parameters)
            if function_name == "excel_read_range":
                return self._read_range(parameters)
            if function_name == "excel_search_workbook":
                return self._search(parameters)
            if function_name == "excel_create_workbook":
                return self._create(parameters)
            if function_name == "excel_write_range":
                return self._write_range(parameters)
            if function_name == "excel_append_rows":
                return self._append_rows(parameters)
        except Exception as exc:
            return f"Excel tool error: {exc}"
        return f"Unknown Excel function: {function_name}"
